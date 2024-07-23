#!/usr/bin/env python3
# encoding: utf-8
# coding style: pep8
# ====================================================
#   Copyright (C) 2024  All rights reserved.
#
#   Author        : cxysailor
#   Email         : cxysailor@163.com
#   File Name     : write_data_into_excel.py
#   Last Modified : 2024-06-28 15:24
#   Describe      : 
#
# ====================================================
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill
import numpy as np

wb = Workbook()
ws = wb.active
ws.title = "somedata"

#  生成一些10x10的数据，并写入单元格区域"A1:J10"
for row in range(1, 11):
    for col in range(1, 11):
        ws.cell(row, col).value = np.random.randn()

#  将数据按行求和，并将结果写入每一行数据后面
i = 0
for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10):
    summary = 0
    for cell in row:
        summary += cell.value
    ws.cell(i + 1, 11).value = summary
    i += 1

#  设置颜色和填充
color = Color(rgb="9BC2E8")
fill = PatternFill(patternType="solid", fgColor=color)
#  填充颜色到单元格区域"K1:K10"
for i in range(10):
    ws.cell(i + 1, 11).fill = fill

wb.save(filename="sum.xlsx")
wb.close()
