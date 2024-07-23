#!/usr/bin/env python3
# encoding: utf-8
# coding style: pep8
# ====================================================
#   Copyright (C) 2024  All rights reserved.
#
#   Author        : cxysailor
#   Email         : cxysailor@163.com
#   File Name     : adding_patterns.py
#   Last Modified : 2024-07-04 14:11
#   Describe      : 
#
# ====================================================
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

wb = Workbook()
ws = wb.active

rows = [
    ("Sample",),
    (1,),
    (2,),
    (3,),
    (2,),
    (3,),
    (3,),
    (1,),
    (2,)
]

for r in rows:
    ws.append(r)

c = BarChart()
data = Reference(ws, min_col=1, min_row=1, max_row=8)
c.add_data(data, titles_from_data=True)
c.title = "Chart with pattern"

# set a pattern for the whole series
series = c.series[0]
fill = PatternFillProperties(prst="pct5")
fill.foreground = ColorChoice(prstClr="red")
fill.background = ColorChoice(prstClr="blue")
series.graphicalProperties.pattFill = fill

# set a pttern for a 6th data point (0-indexed)
pt =DataPoint(idx=5)
pt.graphicalProperties.apttFill = PatternFillProperties(prst="ltHorz")
series.dPt.append(pt)

ws.add_chart(c, "C1")

wb.save(filename="pattern.xlsx")
wb.close()
